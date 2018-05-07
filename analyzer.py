from __future__ import print_function
import numpy as np
from PIL import Image
from PIL import ImageFont
from PIL import ImageDraw
from openpyxl import Workbook
from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.drawing import Drawing
from openpyxl.styles import Font
import webbrowser
import os

@property
def anchor(self):
    anchor = self._anchor
    return anchor

Drawing.anchor = anchor

def _anchor(self, anchor):
    self.anchortype = 'twoCell'
    self.drawing._anchor = anchor

openpyxl.drawing.image.Image.anchor = _anchor

#[rgb, threshold/general, threshold/general up, >OR<OR<> for threshold/general, threshold/nucleus-subtractor, imagepath]
#-1 for treshold/nucleus-subtractor means to not use it
#-0 means > than comparison for threshold/general
#-1 means < than comparison for threshold/general
#-2 means not in <> range comparison for threshold/general

additioner = "f"
wkbkname = ""
savewkbkname = ""
workbook = Workbook()

#last valuesSet value is whether threshold scanning is vertical or not
#2nd value in array is special value
valuesSet = [
    [0, [130], -1, 0, -1, 'analyze1.tif', False, [], False], #black/red
    [1, [45, 50, 45], -1, 0, 60, 'analyze2.tif', False,  [0.4, 0.7, 1], False], #green
    [1, [51], -1, 1, -1, 'analyze3.tif', True, [1], False], #gray
    [0, 60, 90, 2, -1, 'analyze4.tif', False, 60, False] #black/gray  #previous values were 70 & 86 for [1, 2] of comparison.
]

#------------------------------------
#tell program which set to use
#0 - black/red, 1 - green, 2 - gray, 3 - black/gray
setKind = 1

if setKind == 0:
    savewkbkname = "redresults_sheet.xlsx"
else:
    savewkbkname = "combinedresults_sheet.xlsx"
    wkbkname = "redresults_sheet.xlsx"
    workbook = load_workbook(wkbkname)

#tell program whether it is in debug mode
debug = False
#------------------------------------

#rgb threshold
rgb = valuesSet[setKind][0]

#Upper bound for cell from background if needed
upperThreshold = valuesSet[setKind][2]

#How to compare against backround
thresholdComparison = valuesSet[setKind][3]

#Differentiates nucleus from cell
ntSubtractor = valuesSet[setKind][4]

#The image name (eg: cells.png)
imageFilePath = valuesSet[setKind][5]
#Importing image
#- make sure the image is in the same folder as the program
image = Image.open(imageFilePath)
width, height = image.size

verticalScanning = valuesSet[setKind][8]

#Differentiates cell from background
if verticalScanning:
    for i in range(0, len(valuesSet[setKind][7])):
        valuesSet[setKind][7][i] = (valuesSet[setKind][7][i] * height)
else:
    for i in range(0, len(valuesSet[setKind][7])):
        valuesSet[setKind][7][i] = (valuesSet[setKind][7][i] * width)
sectionAreaList = valuesSet[setKind][7]
thresholdList = valuesSet[setKind][1]

#Converting image to RGB array
array = np.array(image)

cellPixelCount = 0

#Array to hold point of cell
cellPointArray = []

#Array to hold border points of cell
cellAreaPointsArray = []

#Array to hold crop dimensions of cell
cellCropArray = []

#Array to hold nucleus thresholds
cellNucleusArray = []
cellNucleusPointsArray = []

#Used to create imaginary bounds for checking cell
cellCheckBoundsXLeft = 40
cellCheckBoundsXRight = 40
cellCheckBoundsY = 48

#Holds given cell positions - if given - otherwise automatic detecting is used.
cellUniversalArray = [];

#Used to create imaginary bounds for checking nucleus
nucleusCheckBoundsX = 5
nucleusCheckBoundsY = 8

#Spreadsheet initializtaion
spreadsheet = workbook.active
spreadsheet.title = "Cell Results"

BOLD_FONT = Font(bold=True)
scaler = setKind * 9
spreadsheet.cell(row=scaler+1, column=1, value="Cell Image:").font = BOLD_FONT
spreadsheet.cell(row=scaler+4, column=1, value="Cell Number:").font = BOLD_FONT
spreadsheet.cell(row=scaler+5, column=1, value="Cell Area:").font = BOLD_FONT
spreadsheet.cell(row=scaler+6, column=1, value="Nuclei Count:").font = BOLD_FONT
spreadsheet.cell(row=scaler+7, column=1, value="Cell Width:").font = BOLD_FONT
spreadsheet.cell(row=scaler+8, column=1, value="Cell Height:").font = BOLD_FONT
scaler = 2 * 9
spreadsheet.cell(row=scaler+1, column=1, value="Cell Image:").font = BOLD_FONT
spreadsheet.cell(row=scaler+4, column=1, value="Cell Number:").font = BOLD_FONT
spreadsheet.cell(row=scaler+5, column=1, value="Cell Area:").font = BOLD_FONT
spreadsheet.cell(row=scaler+6, column=1, value="Nuclei Count:").font = BOLD_FONT
spreadsheet.cell(row=scaler+7, column=1, value="Cell Width:").font = BOLD_FONT
spreadsheet.cell(row=scaler+8, column=1, value="Cell Height:").font = BOLD_FONT
scaler = 3 * 9
spreadsheet.cell(row=scaler+1, column=1, value="Cell Image:").font = BOLD_FONT
spreadsheet.cell(row=scaler+4, column=1, value="Cell Number:").font = BOLD_FONT
spreadsheet.cell(row=scaler+5, column=1, value="Cell Area:").font = BOLD_FONT
spreadsheet.cell(row=scaler+6, column=1, value="Nuclei Count:").font = BOLD_FONT
spreadsheet.cell(row=scaler+7, column=1, value="Cell Width:").font = BOLD_FONT
spreadsheet.cell(row=scaler+8, column=1, value="Cell Height:").font = BOLD_FONT
#Stuff to initiate creating html ouptut page.
htmlDocument = open("celldoc.html", "w")
htmlDocument.truncate()
htmlOutput = """
<html>
<head>
<style>
.holder {
    display: block;
    margin-left: auto;
    margin-right: auto;
    width: 100%;
    padding-left: auto;
    padding-right: auto;
}

div.gallery {
    margin: 5px;
    border: 1px solid #ccc;
    float: left;
    width: 20%;
    font-family: Arial;
    padding-left: auto;
    padding-right: auto;
}

.specialImg {
    display: block;
    margin-left: auto;
    margin-right: auto;
    width: 100%;
}

div.gallery:hover {
    border: 1px solid #777;
}

div.gallery img {
    height: 100px;
    width: auto;
}

div.desc {
    padding: 15px;
    text-align: center;
}

h1 {
margin-left: 15%;
}

h3 {
margin-left: 15%;
}
</style>
<script 
<script>

</script>
</head>
<body>
<div style=\"width: 70%; display: block; margin-left: auto; margin-right: auto\">
<h1>Original Image:</h1>
<img src=\"""" + imageFilePath + """\" class=\"specialImg\"/>
<h1>Cell Results Image:</h1>
<h3>The cells are numbered from left to right, going down from each row.</h3>
<img src=\"cellresults_image.png\" class=\"specialImg\"/>
"""

htmlCounter = 0

def differenceC(num1, num2):
    return abs(int(num1)-int(num2))

def addHtmlImageDiv(imagePath, cellNumber, cellArea, cellWidth, cellHeight, nucleiCount, cellX, cellY, nucleusThreshold):
    global htmlCounter
    global htmlOutput
    cellNumber += 1
    description = "<strong>Cell Number: " + str(cellNumber) + "</strong><br>Cell Area: " + str(cellArea) + "px<br>Cell Width: " + str(cellWidth) + "px<br>Cell Height: " + str(cellHeight) + "px<br>Nuclei Count: " + str(nucleiCount) + "<br>Cell X: " + str(cellX) + "<br>Cell Y: " + str(cellY)
    if (debug):
        description += ("<br>Nucleus Threshold: " + str(nucleusThreshold))
    
    if htmlCounter == 0:
        htmlOutput += "<div class=\"holder\">"
    
    htmlOutput += "<div class=\"gallery\"><img src=\"" + imagePath + "\"/><div class\"desc\">" + description + "</div></div>"
    
    if htmlCounter == 3:
        htmlOutput += "</div>"
        htmlCounter = 0
    else:
        htmlCounter += 1

def addHtmlImageDaughterDiv(imagePath, cellNumber, cellArea, cellWidth, cellHeight, cellX, cellY):
    global htmlCounter
    global htmlOutput
    description = "<strong>Cell Number: " + str(cellNumber) + "</strong><br>Cell Area: " + str(cellArea) + "px<br>Cell Width: " + str(cellWidth) + "px<br>Cell Height: " + str(cellHeight) + "px<br>Cell X: " + str(cellX) + "<br>Cell Y: " + str(cellY)
    
    if htmlCounter == 0:
        htmlOutput += "<div class=\"holder\">"
    
    htmlOutput += "<div class=\"gallery\"><img src=\"" + imagePath + "\"/><div class\"desc\">" + description + "</div></div>"
    
    if htmlCounter == 3:
        htmlOutput += "</div>"
        htmlCounter = 0
    else:
        htmlCounter += 1

def addImageToSpreadsheet(imagePath, Row, Column):
    img = openpyxl.drawing.image.Image(imagePath)
    anchor = TwoCellAnchor(_from=AnchorMarker(row=Row-1, col=Column-1), to=AnchorMarker(row=(Row+2), col=(Column)))
    img.anchor(anchor)
    spreadsheet.add_image(img)

#Tests whether the point is not part of an already found cell
def pointIsNew(x, y):
    xIsOkay = False
    yIsOkay = False
    if len(cellUniversalArray) == 0:
        for i in range(len(cellPointArray)):
            xIsOkay = (cellPointArray[i][0] - cellCheckBoundsXLeft <= x) and (cellPointArray[i][0] + cellCheckBoundsXRight >= x)
            yIsOkay = (cellPointArray[i][1] <= y) and (cellPointArray[i][1] + cellCheckBoundsY >= y)
            if ((xIsOkay) and (yIsOkay)):
                return False
    else:
        for i in range(len(cellUniversalArray)):
            xIsOkay = (x >= cellUniversalArray[i][0][0]) and (x <= cellUniversalArray[i][0][0] + cellUniversalArray[i][0][2])
            yIsOkay = (y >= cellUniversalArray[i][0][1]) and (y <= cellUniversalArray[i][0][1] + cellUniversalArray[i][0][3])
            if ((xIsOkay) and (yIsOkay)):
                return False
    return True

#Tests whether the point matches the color of the cell
def pointIsCellAlt(x, y, pastX, pastY):
    #Makes sure the point provided is within the image
    if (valuesSet[setKind][6]):
        try:
            rgbValue = array[y, x, rgb]
            rgbValueComp = array[pastY, pastX, rgb]
            diff = differenceC(rgbValue, rgbValueComp)
            #if (diff < ):
            #    return pointIsCell(x, y)
            #else:
            #/    return True
            return diff > 10
        except IndexError:
            return False
    else:
        for i in range(0, len(sectionAreaList)):
            if (verticalScanning):
                if (y <= sectionAreaList[i]):
                    return pointIsCell(x, y, thresholdList[i])
            elif (x <= sectionAreaList[i]):
                return pointIsCell(x, y, thresholdList[i])
        return pointIsCell(x, y, thresholdList[0])

def redirectPointIsCell(x, y):
    for i in range(0, len(sectionAreaList)):
        if (verticalScanning):
            if (y <= sectionAreaList[i]):
                return pointIsCell(x, y, thresholdList[i])
        elif (x <= sectionAreaList[i]):
            return pointIsCell(x, y, thresholdList[i])
    return pointIsCell(x, y, trhesholdList[0])

def pointIsCell(x, y, threshold):
    #Makes sure the point provided is within the image
    try:
        rgbValue = array[y, x, rgb]
        if thresholdComparison == 0:
            return rgbValue > threshold
        elif thresholdComparison == 1:
            return rgbValue < threshold
        elif thresholdComparison == 2:
            return rgbValue < threshold or rgbValue > upperThreshold
        else:
            print("A threshold comparison was not specified!")
            raise ValueError
        return array[y, x, rgb] > threshold if isGreaterThan else array[y, x, rgb] < threshold
    except IndexError:
        return False

def pointIsNucleus(x, y, cellNumber):
    global cellNucleusArray
    try:
        return array[y, x, rgb] > cellNucleusArray[cellNumber]
    except IndexError:
        return False

def pointIsNewNucleus(x, y, cellNumber):
    global cellNucleusPointsArray
    for i in range(len(cellNucleusPointsArray[cellNumber])):
        xIsOkay = (cellNucleusPointsArray[cellNumber][i][0] - nucleusCheckBoundsX <= x) and (cellNucleusPointsArray[cellNumber][i][0] + nucleusCheckBoundsX >= x)
        yIsOkay = (cellNucleusPointsArray[cellNumber][i][1] <= y) and (cellNucleusPointsArray[cellNumber][i][1] + nucleusCheckBoundsY >= y)
        if ((xIsOkay) and (yIsOkay)):
            return False
    return True

#Finds number of cells in image
# - saved as (x, y) inital point for each cell
def findCellAmount():
    for y in range(len(array)):
        for x in range(len(array[0])):
            #Checks if point is correct color and if it is not within a cell that's found.
            if redirectPointIsCell(x, y) and pointIsNew(x, y):
                cellPointArray.append([x, y])
                cellNucleusPointsArray.append([])

def findCellPixelsArea(x, y):
    cellPixelCount = 0
    leftX = 100000
    rightX = 0
    topY = 100000
    bottomY = 0
    highestRGB = 0

    verticalScanList = list()
    pastX = 0
    pastY = 0
    for a in range(x - cellCheckBoundsXLeft, x + cellCheckBoundsXRight):
        topCellFlag = True
        tempLeft = list()
        tempRight = list()
        for b in range(y, y + cellCheckBoundsY):
            if pointIsCellAlt(a, b, pastX, pastY):
                tempLeft.append(b)
            pastY = b
        for b in range(y + cellCheckBoundsY, y, -1):
            if pointIsCellAlt(a, b, pastX, pastY):
                tempRight.append(b)
        pastX = a
        tempList = list()
        if (len(tempLeft) != 0 and len(tempRight) != 0):
            for yval in range(tempLeft[0], tempRight[0]):
                tempList.append([a, yval])
            if (len(tempList) != 0):
                verticalScanList.append(tempList)

        passer = a
    horizontalScanList = list()

    if (setKind != 3):
        for b in range(y, y + cellCheckBoundsY):
            leftCellFlag = True
            tempLeft = list()
            tempRight = list()
            for a in range(x - cellCheckBoundsXLeft, x + cellCheckBoundsXRight):
                if pointIsCellAlt(a, b, pastX, pastY):
                    tempLeft.append(a)
                pastX = a
            for a in range(x + cellCheckBoundsXRight, x - cellCheckBoundsXLeft, -1):
                if pointIsCellAlt(a, b, pastX, pastY):
                    tempRight.append(a)
                pastX = a
            tempList = list()
            if (len(tempLeft) != 0 and len(tempRight) != 0):
                for xval in range(tempLeft[0], tempRight[0]):
                    tempList.append([xval, b])
                horizontalScanList.append(tempList)
            pastY = b
        allPointsArray = list()
        horizontalAllPointsArray = list()
        for columnList in verticalScanList:
            tempArray = list()
            for coords in columnList:
                for rowList in horizontalScanList:
                    if coords in rowList:
                        tempArray.append(coords)
                        if coords[0] < leftX:
                            leftX = coords[0]
                        if coords[0] > rightX:
                            rightX = coords[0]
                        if coords[1] < topY:
                            topY = coords[1]
                        if coords[1] > bottomY:
                            bottomY = coords[1]
                        break
            if (len(tempArray) != 0):
                allPointsArray.append(tempArray)

        for rowList in horizontalScanList:
            tempArray = list()
            for coords in rowList:
                for columnList in verticalScanList:
                    if coords in columnList:
                        tempArray.append(coords)
                        break
            if (len(tempArray) != 0):
                horizontalAllPointsArray.append(tempArray)

        for verticalListOfPoints in allPointsArray:
            cellPixelCount += len(verticalListOfPoints)
            cellAreaPointsArray.append(verticalListOfPoints[0])
            cellAreaPointsArray.append(verticalListOfPoints[len(verticalListOfPoints)-1])
            for vPoint in verticalListOfPoints:
                if array[vPoint[1], vPoint[0], rgb] > highestRGB:
                    try:
                        highestRGB = array[b, a, rgb]
                    except IndexError:
                        pass

        for horizontalListOfPoints in horizontalAllPointsArray:
            cellAreaPointsArray.append(horizontalListOfPoints[0])
            cellAreaPointsArray.append(horizontalListOfPoints[len(horizontalListOfPoints)-1])

        leftVal = leftX-7
        width, height = image.size
        if (leftVal < 0 or leftVal > 99990):
            leftVal = 0

        topVal = topY-5
        if (topVal < 0 or topVal > 99990):
            topVal = 0

        rightVal = rightX+5
        if (rightVal > width):
            rightVal = width
        
        bottomVal = bottomY+6
        if (bottomVal > height):
            bottomVal = height

        cellCropArray.append([leftVal, topVal, rightVal, bottomVal])
        cellNucleusArray.append((highestRGB-20) - ((highestRGB - thresholdList[0])/2.75));
        return cellPixelCount
    else:
        for verticalListOfPoints in verticalScanList:
            cellPixelCount += len(verticalListOfPoints)
            print(verticalListOfPoints[0])
            cellAreaPointsArray.append(verticalListOfPoints[0])
            cellAreaPointsArray.append(verticalListOfPoints[len(verticalListOfPoints)-1])
            for vPoint in verticalListOfPoints:
                if vPoint[0] < leftX:
                    leftX = vPoint[0]
                if vPoint[0] > rightX:
                    rightX = vPoint[0]
                if vPoint[1] < topY:
                    topY = vPoint[1]
                if vPoint[1] > bottomY:
                    bottomY = vPoint[1]
                if array[vPoint[1], vPoint[0], rgb] > highestRGB:
                    highestRGB = array[b, a, rgb]
        leftVal = leftX-7
        width, height = image.size
        if (leftVal < 0):
            leftVal = 0

        topVal = topY-5
        if (topVal < 0):
            topVal = 0

        rightVal = rightX+5
        if (rightVal > width):
            rightVal = width
        
        bottomVal = bottomY+6
        if (bottomVal > height):
            bottomVal = height

        cellCropArray.append([leftVal, topVal, rightVal, bottomVal])
        cellNucleusArray.append((highestRGB-20) - ((highestRGB - thresholdList[0])/2.75));
        return cellPixelCount

def findCellPixelsAreaGivenInput(cellNum):
    cellPixelCount = 0
    leftX = 100000
    rightX = 0
    topY = 100000
    bottomY = 0
    highestRGB = 0

    verticalScanList = list()
    pastX = 0
    pastY = 0
    for a in range(cellUniversalArray[cellNum][0][0], cellUniversalArray[cellNum][0][0] + cellUniversalArray[cellNum][0][2]):
        topCellFlag = True
        tempLeft = list()
        tempRight = list()
        for b in range(cellUniversalArray[cellNum][0][1], cellUniversalArray[cellNum][0][1] + cellUniversalArray[cellNum][0][3]):
            if pointIsCellAlt(a, b, pastX, pastY):
                tempLeft.append(b)
            pastY = b
        for b in range(cellUniversalArray[cellNum][0][1] + cellUniversalArray[cellNum][0][3], cellUniversalArray[cellNum][0][1], -1):
            if pointIsCellAlt(a, b, pastX, pastY):
                tempRight.append(b)
        pastX = a
        tempList = list()
        if (len(tempLeft) != 0 and len(tempRight) != 0):
            for yval in range(tempLeft[0], tempRight[0]):
                tempList.append([a, yval])
            if (len(tempList) != 0):
                verticalScanList.append(tempList)

        passer = a
    horizontalScanList = list()

    if (setKind != 3):
        for b in range(cellUniversalArray[cellNum][0][1], cellUniversalArray[cellNum][0][1] + cellUniversalArray[cellNum][0][3]):
            leftCellFlag = True
            tempLeft = list()
            tempRight = list()
            for a in range(cellUniversalArray[cellNum][0][0], cellUniversalArray[cellNum][0][0] + cellUniversalArray[cellNum][0][2]):
                if pointIsCellAlt(a, b, pastX, pastY):
                    tempLeft.append(a)
                pastX = a
            for a in range(cellUniversalArray[cellNum][0][0] + cellUniversalArray[cellNum][0][2], cellUniversalArray[cellNum][0][0], -1):
                if pointIsCellAlt(a, b, pastX, pastY):
                    tempRight.append(a)
                pastX = a
            tempList = list()
            if (len(tempLeft) != 0 and len(tempRight) != 0):
                for xval in range(tempLeft[0], tempRight[0]):
                    tempList.append([xval, b])
                horizontalScanList.append(tempList)
            pastY = b
        allPointsArray = list()
        horizontalAllPointsArray = list()
        for columnList in verticalScanList:
            tempArray = list()
            for coords in columnList:
                for rowList in horizontalScanList:
                    if coords in rowList:
                        tempArray.append(coords)
                        if coords[0] < leftX:
                            leftX = coords[0]
                        if coords[0] > rightX:
                            rightX = coords[0]
                        if coords[1] < topY:
                            topY = coords[1]
                        if coords[1] > bottomY:
                            bottomY = coords[1]
                        break
            if (len(tempArray) != 0):
                allPointsArray.append(tempArray)

        for rowList in horizontalScanList:
            tempArray = list()
            for coords in rowList:
                for columnList in verticalScanList:
                    if coords in columnList:
                        tempArray.append(coords)
                        break
            if (len(tempArray) != 0):
                horizontalAllPointsArray.append(tempArray)

        for verticalListOfPoints in allPointsArray:
            cellPixelCount += len(verticalListOfPoints)
            cellAreaPointsArray.append(verticalListOfPoints[0])
            cellAreaPointsArray.append(verticalListOfPoints[len(verticalListOfPoints)-1])
            for vPoint in verticalListOfPoints:
                if array[vPoint[1], vPoint[0], rgb] > highestRGB:
                    highestRGB = array[b, a, rgb]

        #debug
        #for verticalListOfPoints in allPointsArray:
        #    for vList in verticalListOfPoints:
        #        cellAreaPointsArray.append(vList)
        #for horizontalListOfPoints in horizontalAllPointsArray:
        #    for hList in horizontalListOfPoints:
        #        cellAreaPointsArray.append(hList)
        #debug

        for horizontalListOfPoints in horizontalAllPointsArray:
            cellAreaPointsArray.append(horizontalListOfPoints[0])
            cellAreaPointsArray.append(horizontalListOfPoints[len(horizontalListOfPoints)-1])

        leftVal = leftX-7
        width, height = image.size
        if (leftVal < 0):
            leftVal = 0

        topVal = topY-5
        if (topVal < 0):
            topVal = 0

        rightVal = rightX+5
        if (rightVal > width):
            rightVal = width
        
        bottomVal = bottomY+6
        if (bottomVal > height):
            bottomVal = height
            
        cellCropArray.append([leftVal, topVal, rightVal, bottomVal])
        cellNucleusArray.append((highestRGB-20) - ((highestRGB - thresholdList[0])/2.75));
        return cellPixelCount
    else:
        for verticalListOfPoints in verticalScanList:
            cellPixelCount += len(verticalListOfPoints)
            #for vList in verticalListOfPoints:
            #    cellAreaPointsArray.append(vList)
            print(verticalListOfPoints[0])
            cellAreaPointsArray.append(verticalListOfPoints[0])
            cellAreaPointsArray.append(verticalListOfPoints[len(verticalListOfPoints)-1])
            for vPoint in verticalListOfPoints:
                if vPoint[0] < leftX:
                    leftX = vPoint[0]
                if vPoint[0] > rightX:
                    rightX = vPoint[0]
                if vPoint[1] < topY:
                    topY = vPoint[1]
                if vPoint[1] > bottomY:
                    bottomY = vPoint[1]
                if array[vPoint[1], vPoint[0], rgb] > highestRGB:
                    highestRGB = array[b, a, rgb]
        cellCropArray.append([leftX-7, topY-5, rightX+5, bottomY+6])
        cellNucleusArray.append((highestRGB-20) - ((highestRGB - thresholdList[0])/2.75));
        return cellPixelCount

def getNucleusCount(cellNumber):
    global cellNucleusArray
    nucleusCount = 100
    while nucleusCount > 10:
        nucleusCount = 0
        for y in range(cellCropArray[cellNumber][1], cellCropArray[cellNumber][3]):
            passer = y
            for x in range(cellCropArray[cellNumber][0], cellCropArray[cellNumber][2]):
                hi = x
                if pointIsNucleus(x, y, cellNumber) and pointIsNewNucleus(x, y, cellNumber):
                    cellNucleusPointsArray[cellNumber].append([x, y])
                    nucleusCount += 1
                hi = x
            passer = y
        if (nucleusCount > 12):
            del cellNucleusPointsArray[cellNumber][:]
            cellNucleusArray[cellNumber] += 15
    return nucleusCount

def getNucleusCountu(cellNumber):
    global cellNucleusArray
    nucleusCount = 100
    while nucleusCount > 10:
        nucleusCount = 0
        for y in range(cellCropArray[cellNumber][1], cellCropArray[cellNumber][3]):
            passer = y
            for x in range(cellCropArray[cellNumber][0], cellCropArray[cellNumber][2]):
                hi = x
                if pointIsNucleus(x, y, cellNumber) and pointIsNewNucleus(x, y, cellNumber):
                    cellNucleusPointsArray[cellNumber].append([x, y])
                    nucleusCount += 1
                hi = x
            passer = y
        if (nucleusCount > 12):
            del cellNucleusPointsArray[cellNumber][:]
            cellNucleusArray[cellNumber] += 15
    return nucleusCount

tempString = raw_input("Do you have output from the Cell Tool? (Put 'none' for automation, and make sure there are no spaces in your input) - ");
if tempString != "none":
    cellUniversalArray = eval(tempString)
    for i in range(len(cellUniversalArray)):
        cellNucleusPointsArray.append([])

#Keeps running until correct bounds are used to find the correct amount of cells
if (len(cellUniversalArray) == 0):
    while (True):
        del cellPointArray[:]
        findCellAmount();
        print (cellPointArray)
        print (str(len(cellPointArray)) + " cells found in the image.")
        response = int(input("Is that right? (1-Yes / 0-No)"))
        if (response == 1):
            break
        else:
            xLeftChange = int(input("Change cell check bounds left X by: "))
            xRightChange = int(input("Change cell check bounds right X by: "))
            yChange = int(input("Change cell check bounds Y by: "))
            cellCheckBoundsXLeft += xLeftChange
            cellCheckBoundsXRight += xRightChange
            cellCheckBoundsY += yChange
            print ("-----------------------")
            print (str(cellCheckBoundsXLeft))
            print (str(cellCheckBoundsXRight))
            print (str(cellCheckBoundsY))
        
print ("&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&&");
print ("The cells are numbered from left to right, going down from each row.");

#Preparing program to write numbers on top of image
font = ImageFont.truetype("timesnewroman.ttf", 18)
draw = ImageDraw.Draw(image)

borderImg = Image.open("whiteborder.png")
nucleusBorderImg = Image.open("whitenucleusborder.png")

cellAreaArray = list()
cellNucleusCountArray = list()

if (len(cellUniversalArray) == 0):
    for i in range(len(cellPointArray)):
        #Gets area of the cell
        cellAreaArray.append(findCellPixelsArea(cellPointArray[i][0], cellPointArray[i][1]))
        if (valuesSet[setKind][4] != -1):
            cellNucleusCountArray.append(getNucleusCount(i))
        else:
            cellNucleusCountArray.append("N/A")
        if (debug):
            for a in cellNucleusPointsArray[i]:
                image.paste(nucleusBorderImg, (a[0]-5, a[1]), nucleusBorderImg)
else:
    for i in range(len(cellUniversalArray)):
        #Gets area of the cell
        cellAreaArray.append(findCellPixelsAreaGivenInput(i))
        if (valuesSet[setKind][4] != -1):
            cellNucleusCountArray.append(getNucleusCount(i))
        else:
            cellNucleusCountArray.append("N/A")
        if (debug):
            for a in cellNucleusPointsArray[i]:
                image.paste(nucleusBorderImg, (a[0]-5, a[1]), nucleusBorderImg)

#Saving image to folder
draw = ImageDraw.Draw(image)
pixels = image.load()
for point in cellAreaPointsArray:
    try:
        #print("X: " + str(point[0]) + " Y: " + sgtr(point[1]))
        pixels[point[0], point[1]] = (255, 255, 255)
    except IndexError:
        pass


grayImage = Image.open(valuesSet[2][5])
blackImage = Image.open(valuesSet[3][5]) 

if len(cellUniversalArray) == 0:
    for i in range(len(cellAreaArray)-1, 0, -1):
        if (cellAreaArray[i] == 0):
            cellAreaArray.pop(i)
            cellNucleusCountArray.pop(i)
            cellCropArray.pop(i)
            cellPointArray.pop(i)
    print(cellCropArray)
    print("[", end='')
    for cropList in cellCropArray:
        print("[" + str(cropList[0]) + ", " + str(cropList[1]) + ", " + str(cropList[2]-cropList[0]) + ", " + str(cropList[3]-cropList[1]) + "],", end='')
    print("]\n\n")
    for cropList in cellCropArray:
        print("[[" + str(cropList[0]) + ", " + str(cropList[1]) + ", " + str(cropList[2]-cropList[0]) + ", " + str(cropList[3]-cropList[1]) + "]],", end='')
    print("]")
    for i in range(len(cellPointArray)):
        cellCrop = image.crop((cellCropArray[i][0], cellCropArray[i][1], cellCropArray[i][2], cellCropArray[i][3]))
        cellCrop.save("cell" + str(i) + str(setKind) + ".png")
        addHtmlImageDiv("cell" + str(i) + str(setKind) + ".png", i, cellAreaArray[i], cellCropArray[i][2]-cellCropArray[i][0], cellCropArray[i][3]-cellCropArray[i][1], cellNucleusCountArray[i], cellPointArray[i][0], cellPointArray[i][1], cellNucleusArray[i])
        scaler = setKind * 9
        columnVal = (i+2)*3-2
        spreadsheet.cell(row=(scaler+4), column=columnVal, value=(i+1))
        spreadsheet.cell(row=(scaler+5), column=columnVal, value=cellAreaArray[i])
        spreadsheet.cell(row=(scaler+6), column=columnVal, value=cellNucleusCountArray[i])
        spreadsheet.cell(row=(scaler+7), column=columnVal, value=(cellCropArray[i][2]-cellCropArray[i][0]))
        spreadsheet.cell(row=(scaler+8), column=columnVal, value=(cellCropArray[i][3]-cellCropArray[i][1]))
        
        grayImage.crop((cellCropArray[i][0], cellCropArray[i][1], cellCropArray[i][2], cellCropArray[i][3])).save("cell" + str(i) + "2.png")
        scaler = 2 * 9
        spreadsheet.cell(row=(scaler+4), column=columnVal, value=(i+1))
        spreadsheet.cell(row=(scaler+5), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+6), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+7), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+8), column=columnVal, value=-1)
        
        blackImage.crop((cellCropArray[i][0], cellCropArray[i][1], cellCropArray[i][2], cellCropArray[i][3]+7)).save("cell" + str(i) + "3.png")
        scaler = 3 * 9
        spreadsheet.cell(row=(scaler+4), column=columnVal, value=(i+1))
        spreadsheet.cell(row=(scaler+5), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+6), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+7), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+8), column=columnVal, value=-1)

    for i in range(len(cellPointArray)):
        #Overlays number on top of image
        draw.text((cellPointArray[i][0]-20, cellPointArray[i][1]), str(i+1), (255, 255, 255), font=font)
        if (debug):
            image.paste(borderImg, (cellPointArray[i][0]-40, cellPointArray[i][1]), borderImg)
else:
    print(cellCropArray)
    for i in range(len(cellUniversalArray)):
        valSet = (0, 0, 0, 0)
        try:
            cellCrop = image.crop((cellCropArray[i][0], cellCropArray[i][1], cellCropArray[i][2], cellCropArray[i][3]))
            valSet = (cellCropArray[i][0], cellCropArray[i][1], cellCropArray[i][2], cellCropArray[i][3])
            cellCrop.save("cell" + str(i) + str(setKind) + ".png")
        except SystemError:
            cellCrop = image.crop((cellUniversalArray[i][0][0], cellUniversalArray[i][0][1], cellUniversalArray[i][0][0]+cellUniversalArray[i][0][2], cellUniversalArray[i][0][1]+cellUniversalArray[i][0][3]))
            valSet = (cellUniversalArray[i][0][0], cellUniversalArray[i][0][1], cellUniversalArray[i][0][0]+cellUniversalArray[i][0][2], cellUniversalArray[i][0][1]+cellUniversalArray[i][0][3])
            cellCrop.save("cell" + str(i) + str(setKind) + ".png")
        addHtmlImageDiv("cell" + str(i) + str(setKind) + ".png", i, cellAreaArray[i], cellCropArray[i][2]-cellCropArray[i][0], cellCropArray[i][3]-cellCropArray[i][1], cellNucleusCountArray[i], cellUniversalArray[i][0][0], cellUniversalArray[i][0][1], cellNucleusArray[i])
        columnVal = (i+2)*3-2
        scaler = setKind * 9
        spreadsheet.cell(row=(scaler+4), column=columnVal, value=(i+1))
        spreadsheet.cell(row=(scaler+5), column=columnVal, value=cellAreaArray[i])
        spreadsheet.cell(row=(scaler+6), column=columnVal, value=cellNucleusCountArray[i])
        spreadsheet.cell(row=(scaler+7), column=columnVal, value=(cellCropArray[i][2]-cellCropArray[i][0]))
        spreadsheet.cell(row=(scaler+8), column=columnVal, value=(cellCropArray[i][3]-cellCropArray[i][1]))
        
        valSet = (cellCropArray[i][0]-1, cellCropArray[i][1]-5, cellCropArray[i][2]+5, cellCropArray[i][3]+5)
        grayImage.crop(valSet).save("cell" + str(i) + "2.png")
        scaler = 2 * 9
        spreadsheet.cell(row=(scaler+4), column=columnVal, value=(i+1))
        spreadsheet.cell(row=(scaler+5), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+6), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+7), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+8), column=columnVal, value=-1)
        
        valSet = (cellCropArray[i][0]-10, cellCropArray[i][1]-2, cellCropArray[i][2]-1, cellCropArray[i][3]+8)
        
        blackImage.crop(valSet).save("cell" + str(i) + "3.png")
        scaler = 3 * 9
        spreadsheet.cell(row=(scaler+4), column=columnVal, value=(i+1))
        spreadsheet.cell(row=(scaler+5), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+6), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+7), column=columnVal, value=-1)
        spreadsheet.cell(row=(scaler+8), column=columnVal, value=-1)

    for i in range(len(cellUniversalArray)):
        #Overlays number on top of image
        draw.text((cellUniversalArray[i][0][0], cellUniversalArray[i][0][1]), str(i+1), (255, 255, 255), font=font)
        if (debug):
            image.paste(borderImg, (cellUniversalArray[i][0][0], cellUniversalArray[i][0][1]), borderImg)

if len(cellUniversalArray) == 0:
    for i in range(len(cellPointArray)):
        addImageToSpreadsheet("cell" + str(i) + str(setKind) + ".png", (setKind*9)+1, (i+2)*3-2)
        addImageToSpreadsheet("cell" + str(i) + "2.png", (2*9)+1, (i+2)*3-2)
        addImageToSpreadsheet("cell" + str(i) + "3.png", (3*9)+1, (i+2)*3-2)
else:
    for i in range(len(cellUniversalArray)):
        addImageToSpreadsheet("cell" + str(i) + str(setKind) + ".png", (setKind*9)+1, (i+2)*3-2)
        addImageToSpreadsheet("cell" + str(i) + "2.png", (2*9)+1, (i+2)*3-2)
        addImageToSpreadsheet("cell" + str(i) + "3.png", (3*9)+1, (i+2)*3-2)

incrementer = width/10
for x in range(incrementer, width, incrementer):
    for y in range(len(array)):
        pixels[x, y] = (255, 255, 255)

incrementer = height/10
for y in range(incrementer, height, incrementer):
    for x in range(len(array[0])):
        pixels[x, y] = (255, 255, 255)
#Drawing borders on cell
image.save("cellresults_image.png")


workbook.save(savewkbkname)

print ("Check your folder to see a document called 'celldoc.html' if it doesn't show.")

htmlOutput += "</div></body></html>"
htmlDocument.write(htmlOutput)
htmlDocument.close()





                
